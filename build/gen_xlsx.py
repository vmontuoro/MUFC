import re, datetime as dt, os, sys, subprocess
from itertools import combinations
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.run([sys.executable,"-m","pip","install","openpyxl","--break-system-packages","-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
from dribl_parse import read
from pitch_capacity import category,CATLABEL,conflicts
HERE=os.path.dirname(os.path.abspath(__file__)); OUTDIR=os.path.dirname(HERE)
OUT=os.path.join(OUTDIR,"Manningham_fixtures_and_overlaps.xlsx")
files={"Pettys Reserve":os.path.join(HERE,"raw_pettys.txt"),"Powerful Owl Park":os.path.join(HERE,"raw_powl.txt"),"Timber Ridge Reserve":os.path.join(HERE,"raw_timber.txt"),"Wilsons Rd Reserve":os.path.join(HERE,"raw_wilsons.txt")}
MONTHS={m:i for i,m in enumerate(["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"],1)}
def parse_date(s):
    p=s.split(); return dt.date(int(p[3]),MONTHS[p[2][:3]],int(p[1]))
def duration(t,comp):
    if "All Abilities" in comp: return 25
    if "MiniRoos" in t or re.search(r'U0[6-9]\b',t) or "U10" in t or "U11" in t: return 40
    if "Seniors" in t or "Reserves" in t: return 90
    if re.search(r'U2[013]\b',t): return 90
    if re.search(r'U1[78]\b',t) or "17/18" in t: return 70
    if re.search(r'U1[56]\b',t): return 70
    if re.search(r'U1[234]\b',t): return 60
    return 60
def age_token(t):
    if "17/18" in t: return "U17/18"
    m=re.search(r'(U\d{2}|Seniors|Reserves)',t); return m.group(1) if m else ""
def field_info(pitch,ground):
    p=pitch.replace(ground,"").strip()
    m=re.search(r'Pitch\s+(\d+)',p); num=m.group(1) if m else "?"
    after=p[m.end():].lstrip() if m else ""
    letter=after[0] if after and after[0].isalpha() and not after.startswith("(") else ""
    return num,("Sub" if (letter or "Midi" in p) else "Full"),p
def fmt(m): return f"{m//60:02d}:{m%60:02d}"
import manual_games
games=[]
for ground,path in files.items():
    _tuples=read(path)
    if ground=="Pettys Reserve": _tuples=list(_tuples)+manual_games.build(_tuples,parse_date)
    for date_s,time_s,home,away,comp,pitch,rnd in _tuples:
        d=parse_date(date_s); hh,mm=map(int,time_s.split(":")); start=hh*60+mm
        dur=duration(home+" "+away,comp); num,ftype,plabel=field_info(pitch,ground)
        cp=[x.strip() for x in comp.split("|")]
        _age=age_token(home if "Manningham" in home else away+" "+home)
        games.append(dict(date=d,day=date_s.split()[0],time=time_s,start=start,end=start+dur,dur=dur,ground=ground,
            field=num,ftype=ftype,pitch=plabel,age=_age,cat=category(_age,cp[0]),
            home=home,away=away,comp=cp[0],grade=cp[1] if len(cp)>1 else "",rnd=rnd.replace("Round ","R"),clash=""))
import overrides as _ovr
def gkey_of(g):
    """Must match gen_clashes.gkey_of exactly: iso|ground|pitch|time|home"""
    return "|".join([g["date"].isoformat(),g["ground"],g["pitch"],g["time"],g["home"]])
# Apply saved manual moves BEFORE the overlap check, so the sheet agrees with the
# clash page about where a game is actually being played.
if _ovr.apply(games,gkey_of):
    for g in games:                       # pitch changed, so field number / Full-vs-Sub must follow
        if g.get("override"): g["field"],g["ftype"],_=field_info(g["pitch"],g["ground"])
# Clash = OVER CAPACITY, not merely "two games on one field" — see pitch_capacity.py.
clashes=conflicts(games,lambda g:(g["ground"],g["date"],g["field"]))
for c in clashes:
    for g in c["games"]: g["clash"]="YES"
clashes.sort(key=lambda c:(c["key"][1],c["key"][0],c["key"][2],c["at"]))
peak={}; daydict={}
for g in games: daydict.setdefault((g["ground"],g["date"]),[]).append(g)
for (gr,d),gs in daydict.items():
    evs=[]
    for g in gs: evs+=[(g["start"],1,g),(g["end"],-1,g)]
    evs.sort(key=lambda e:(e[0],e[1])); cur=0; active=[]
    for t,typ,g in evs:
        if typ==1:
            cur+=1; active.append(g)
            if gr not in peak or cur>peak[gr][0]:
                peak[gr]=(cur,d,t,min(x["end"] for x in active),sorted({x["pitch"] for x in active}))
        else:
            cur-=1; active=[x for x in active if x is not g]
games.sort(key=lambda g:(g["ground"],g["date"],g["field"],g["start"]))
wb=Workbook(); FONT="Arial"
hdr_fill=PatternFill("solid",fgColor="1F3864"); hdr_font=Font(name=FONT,bold=True,color="FFFFFF",size=10)
clash_fill=PatternFill("solid",fgColor="FFC7CE"); band=PatternFill("solid",fgColor="F2F2F2")
thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal="center",vertical="center"); left=Alignment(horizontal="left",vertical="center")
fmt2=lambda m:f"{m//60:02d}:{m%60:02d}"
def style_header(ws,ncols,row=1):
    for c in range(1,ncols+1):
        cell=ws.cell(row=row,column=c); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=border
    ws.freeze_panes=ws.cell(row=row+1,column=1)
ws=wb.active; ws.title="All Games"
# "Moved From" goes last so the hard-coded Clash? column index (16) below stays put.
cols=["Date","Day","KO","Est.End","Dur(m)","Ground","Field","Pitch Type","Pitch","Age/Grp","Home","Away","Competition","Grade","Rnd","Clash?","Moved From (pending FV)"]
ws.append(cols)
for g in games:
    mf=g.get("moved_from")
    ws.append([g["date"].strftime("%d %b %Y"),g["day"],g["time"],fmt2(g["end"]),g["dur"],g["ground"],g["field"],
        g["ftype"],g["pitch"],g["age"],g["home"],g["away"],g["comp"],g["grade"],g["rnd"],g["clash"],
        f'{mf["ground"]} {mf["pitch"]}' if mf else ""])
style_header(ws,len(cols))
for i,w in enumerate([12,5,7,8,7,17,6,10,27,8,42,42,34,26,5,7,34],1): ws.column_dimensions[get_column_letter(i)].width=w
for r in range(2,ws.max_row+1):
    cl=ws.cell(row=r,column=16).value=="YES"
    for c in range(1,len(cols)+1):
        cell=ws.cell(row=r,column=c); cell.font=Font(name=FONT,size=9,bold=cl and c==16,color="9C0006" if cl else "000000")
        cell.border=border; cell.alignment=center if c in (1,2,3,4,5,6,7,8,10,14,15,16) else left
        if cl: cell.fill=clash_fill
        elif r%2==0: cell.fill=band
ws.auto_filter.ref=f"A1:{get_column_letter(len(cols))}{ws.max_row}"
ws2=wb.create_sheet("Overlaps & Peak Load")
intro=["Overlap & clash check (pitch capacity)",
 "A full pitch is worth 1.0.  U14+/Seniors = 1.0 and must have the pitch alone;  U10-U13 = 0.5;  U6-U9 = 0.25;  All-Abilities = 0.5.",
 "A CLASH = the games running at the same moment on one physical field add up to MORE than 1.0. Small-sided games sharing a field is NOT a clash.",
 "Same rules as the clash-check web page, so the two always agree.",
 "Durations ESTIMATED: MiniRoos 40m, All-Abilities 25m, U12-14 60m, U15-18 70m, Seniors/U20-23 90m."]
for line in intro: ws2.append([line])
ws2["A1"].font=Font(name=FONT,bold=True,size=13,color="1F3864")
for r in range(2,len(intro)+1): ws2.cell(row=r,column=1).font=Font(name=FONT,size=9,italic=True,color="595959")
ws2.append([])
ws2.append([f"A.  Over-capacity moments: {len(clashes)} found"])
ws2.cell(row=ws2.max_row,column=1).font=Font(name=FONT,bold=True,size=11,color="9C0006" if clashes else "375623")
if clashes:
    ws2.append(["#","Date","Ground","Field","From","Until","Rule broken","Load","Games in that window"]); hrow=ws2.max_row
    for n,c in enumerate(clashes,1):
        gr,d,fl=c["key"]
        ws2.append([n,d.strftime("%d %b %Y"),gr,fl,fmt2(c["at"]),fmt2(c["until"]),c["rule"],
            f'{c["mix"]} = {c["units"]:g}',
            "  |  ".join(f'{g["time"]} {g["age"]} {g["pitch"]} ({g["home"]} v {g["away"]})'
                         for g in sorted(c["games"],key=lambda x:x["start"]))])
    style_header(ws2,9,row=hrow)
    for r in range(hrow+1,ws2.max_row+1):
        for c in range(1,10):
            cell=ws2.cell(row=r,column=c); cell.font=Font(name=FONT,size=9); cell.border=border
            cell.alignment=center if c in (1,2,3,4,6,8,9) else left
            if r%2==0: cell.fill=band
else:
    ws2.append(["None. Where games share a physical field, the age groups fit within the 1.0 limit."])
    ws2.cell(row=ws2.max_row,column=1).font=Font(name=FONT,size=9,italic=True)
ws2.append([]); ws2.append([])
ws2.append(["B.  Peak simultaneous games per ground (busiest moment)"])
ws2.cell(row=ws2.max_row,column=1).font=Font(name=FONT,bold=True,size=11,color="1F3864")
ws2.append(["Ground","Max at once","Date","From","Until","Pitches in use then"]); brow=ws2.max_row
for gr in ["Pettys Reserve","Powerful Owl Park","Timber Ridge Reserve"]:
    if gr in peak:
        c,d,s,e,pits=peak[gr]; ws2.append([gr,c,d.strftime("%a %d %b %Y"),fmt2(s),fmt2(e),", ".join(pits)])
style_header(ws2,6,row=brow)
for r in range(brow+1,ws2.max_row+1):
    for c in range(1,7):
        cell=ws2.cell(row=r,column=c); cell.font=Font(name=FONT,size=9); cell.border=border
        cell.alignment=center if c in (2,3,4,5) else left
        if r%2==0: cell.fill=band
for i,w in enumerate([18,13,17,9,9,30,30,26,84],1): ws2.column_dimensions[get_column_letter(i)].width=w
ws3=wb.create_sheet("Summary",0)
ws3.append(["Manningham United Blues - Home Fixtures & Overlap Check"])
ws3.append(["Grounds: Pettys Reserve, Powerful Owl Park, Timber Ridge Reserve"])
ws3.append([f"Source: fv.dribl.com, 2026 season, generated {dt.date.today().strftime('%d %b %Y')}"])
ws3.append([]); ws3.append(["Ground","Games","Date range"])
for gr in ["Pettys Reserve","Powerful Owl Park","Timber Ridge Reserve"]:
    gg=[g for g in games if g["ground"]==gr]
    dr=f'{min(x["date"] for x in gg).strftime("%d %b")} - {max(x["date"] for x in gg).strftime("%d %b %Y")}' if gg else "-"
    ws3.append([gr,len(gg),dr])
ws3.append(["TOTAL",len(games),""]); ws3.append([])
ws3.append(["Over-capacity moments flagged",len(clashes)])
ws3["A1"].font=Font(name=FONT,bold=True,size=14,color="1F3864")
for r in (2,3): ws3.cell(row=r,column=1).font=Font(name=FONT,size=9,italic=True,color="595959")
for c in range(1,4):
    cell=ws3.cell(row=5,column=c); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=border
for r in range(6,10):
    for c in range(1,4):
        cell=ws3.cell(row=r,column=c); cell.border=border; cell.font=Font(name=FONT,size=10,bold=(ws3.cell(row=r,column=1).value=="TOTAL"))
        cell.alignment=center if c>1 else left
ws3.cell(row=11,column=1).font=Font(name=FONT,bold=True,size=11,color="375623")
ws3.cell(row=11,column=2).font=Font(name=FONT,bold=True,size=11,color="375623")
ws3.column_dimensions["A"].width=44; ws3.column_dimensions["B"].width=10; ws3.column_dimensions["C"].width=22
wb.save(OUT)
print("xlsx:",len(games),"games,",len(clashes),"clashes ->",OUT)
